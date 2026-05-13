from __future__ import annotations

import csv
import json
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


@dataclass(frozen=True)
class Document:
    path: str
    text: str
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class Chunk:
    id: str
    document_path: str
    text: str
    start_line: int
    end_line: int
    metadata: dict[str, Any] = field(default_factory=dict)


def load_document(path: str | Path) -> Document:
    """Load simple local document formats without optional dependencies."""

    p = Path(path)
    suffix = p.suffix.lower()
    if suffix in {".txt", ".md"}:
        return Document(str(p), p.read_text(encoding="utf-8"), {"format": suffix[1:]})
    if suffix == ".json":
        data = json.loads(p.read_text(encoding="utf-8"))
        return Document(str(p), json.dumps(data, indent=2), {"format": "json", "data": data})
    if suffix in {".csv", ".tsv", ".xlsx", ".reqif", ".reqifz", ".xml"}:
        rows = load_table_rows(p)
        text = "\n".join(", ".join(f"{k}: {v}" for k, v in row.items()) for row in rows)
        return Document(str(p), text, {"format": suffix[1:], "rows": rows})
    raise ValueError(
        f"Unsupported file type '{suffix}'. Supported: txt, md, csv, tsv, xlsx, json, reqif, reqifz, xml."
    )


def load_csv_rows(path: str | Path) -> list[dict[str, str]]:
    return load_delimited_rows(path, delimiter=",")


def load_table_rows(path: str | Path) -> list[dict[str, str]]:
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return load_delimited_rows(p, delimiter=",")
    if suffix == ".tsv":
        return load_delimited_rows(p, delimiter="\t")
    if suffix == ".xlsx":
        return load_xlsx_rows(p)
    if suffix in {".reqif", ".xml"}:
        return load_reqif_rows(p)
    if suffix == ".reqifz":
        return load_reqifz_rows(p)
    raise ValueError(f"Unsupported table file type '{suffix}'.")


def load_delimited_rows(path: str | Path, delimiter: str) -> list[dict[str, str]]:
    with Path(path).open(newline="", encoding="utf-8-sig") as f:
        return [_clean_row(row) for row in csv.DictReader(f, delimiter=delimiter)]


def load_xlsx_rows(path: str | Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel ingestion requires openpyxl. Install with: pip install -e \".[ingestion]\""
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows: list[dict[str, str]] = []
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            try:
                header = next(iterator)
            except StopIteration:
                continue
            keys = [_normalize_header(value, fallback=f"column_{i}") for i, value in enumerate(header, start=1)]
            if not any(keys):
                continue
            for row_index, values in enumerate(iterator, start=2):
                row = {
                    key: _cell_to_text(value)
                    for key, value in zip(keys, values)
                    if key
                }
                if any(row.values()):
                    row["_sheet"] = sheet.title
                    row["_row"] = str(row_index)
                    rows.append(row)
        return rows
    finally:
        workbook.close()


def load_reqifz_rows(path: str | Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as archive:
        candidates = [name for name in archive.namelist() if name.lower().endswith((".reqif", ".xml"))]
        if not candidates:
            raise ValueError("REQIFZ archive did not contain a .reqif or .xml file.")
        with archive.open(candidates[0]) as f:
            return _parse_reqif_xml(f.read())


def load_reqif_rows(path: str | Path) -> list[dict[str, str]]:
    return _parse_reqif_xml(Path(path).read_bytes())


def _parse_reqif_xml(data: bytes) -> list[dict[str, str]]:
    root = ElementTree.fromstring(data)
    rows: list[dict[str, str]] = []
    for spec_object in root.findall(".//{*}SPEC-OBJECT"):
        row: dict[str, str] = {
            "id": spec_object.attrib.get("IDENTIFIER", ""),
            "long_name": spec_object.attrib.get("LONG-NAME", ""),
        }
        for attr in spec_object.findall(".//{*}ATTRIBUTE-VALUE-STRING"):
            definition = attr.find(".//{*}ATTRIBUTE-DEFINITION-STRING-REF")
            key = definition.text if definition is not None and definition.text else "text"
            row[_normalize_header(key)] = attr.attrib.get("THE-VALUE", "")
        for attr in spec_object.findall(".//{*}ATTRIBUTE-VALUE-XHTML"):
            definition = attr.find(".//{*}ATTRIBUTE-DEFINITION-XHTML-REF")
            key = definition.text if definition is not None and definition.text else "text"
            values = [text.strip() for text in attr.itertext() if text and text.strip()]
            row[_normalize_header(key)] = " ".join(values)
        if any(row.values()):
            rows.append(_clean_row(row))
    return rows


def _clean_row(row: dict[str, Any]) -> dict[str, str]:
    return {_normalize_header(key): _cell_to_text(value) for key, value in row.items()}


def _normalize_header(value: Any, fallback: str = "") -> str:
    text = _cell_to_text(value).strip().lower()
    return text.replace(" ", "_").replace("-", "_") if text else fallback


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def write_json(path: str | Path, data: Any) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(data, indent=2), encoding="utf-8")


def chunk_document(document: Document, max_lines: int = 40, overlap: int = 5) -> list[Chunk]:
    lines = document.text.splitlines()
    chunks: list[Chunk] = []
    if not lines:
        return chunks
    step = max(1, max_lines - overlap)
    for start in range(0, len(lines), step):
        end = min(len(lines), start + max_lines)
        chunk_lines = lines[start:end]
        chunks.append(
            Chunk(
                id=f"{Path(document.path).stem}:{start + 1}-{end}",
                document_path=document.path,
                text="\n".join(chunk_lines),
                start_line=start + 1,
                end_line=end,
            )
        )
        if end == len(lines):
            break
    return chunks
